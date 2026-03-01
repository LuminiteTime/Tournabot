"""Генерация Excel-файла с результатами турнира."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_results_excel(
    data: dict,
    all_rankings: list[list[dict]],
    overall: list[dict],
) -> bytes:
    """
    Создать .xlsx с таблицами матчей + столбцы рейтинга + лист общего зачёта.

    Возвращает байтовое содержимое файла.
    """
    wb = Workbook()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)
    diag_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ── Листы таблиц ───────────────────────────────────────────────────

    for t_idx, table in enumerate(data["tables"]):
        ws = wb.active if t_idx == 0 else wb.create_sheet()
        ws.title = f"Таблица {t_idx + 1}"

        size = table["size"]
        players = table["players"]
        matches = table["matches"]
        rankings = all_rankings[t_idx]

        # Индекс для быстрого доступа: позиция → рейтинг
        rank_by_pos: dict[int, dict] = {r["pos"]: r for r in rankings}

        # Колонки: №, Имя, [1..size], Очки, Соотн., Место
        stats_start_col = size + 3  # первая колонка статистики (после матчей)

        # ── Заголовки ──
        ws.cell(row=1, column=1, value="№").font = header_font
        ws.cell(row=1, column=2, value="Имя").font = header_font
        for j in range(1, size + 1):
            c = ws.cell(row=1, column=j + 2, value=j)
            c.font = header_font
            c.alignment = center
            c.border = border

        for col_offset, label in enumerate(["Очки", "Соотн.", "Место"]):
            c = ws.cell(row=1, column=stats_start_col + col_offset, value=label)
            c.font = header_font
            c.alignment = center
            c.border = border

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 22
        for j in range(1, size + 1):
            ws.column_dimensions[get_column_letter(j + 2)].width = 10
        ws.column_dimensions[get_column_letter(stats_start_col)].width = 8
        ws.column_dimensions[get_column_letter(stats_start_col + 1)].width = 10
        ws.column_dimensions[get_column_letter(stats_start_col + 2)].width = 8

        # ── Строки с результатами ──
        for i in range(1, size + 1):
            row_num = i + 1
            ws.cell(row=row_num, column=1, value=i).font = header_font
            ws.cell(row=row_num, column=2, value=players[i - 1]).font = Font(bold=True)

            for j in range(1, size + 1):
                cell = ws.cell(row=row_num, column=j + 2)
                cell.alignment = center
                cell.border = border

                if i == j:
                    cell.value = "×"
                    cell.fill = diag_fill
                elif i < j:
                    # Верхний правый треугольник (строка i, столбец j)
                    key = f"{i}_{j}"
                    m = matches.get(key, {})
                    if m.get("status") == "finished":
                        s1, s2 = m["score1"], m["score2"]
                        cell.value = f"{s1}:{s2}"
                    else:
                        cell.value = "—"
                else:
                    # Нижний левый треугольник (зеркало)
                    key = f"{j}_{i}"
                    m = matches.get(key, {})
                    if m.get("status") == "finished":
                        s1, s2 = m["score1"], m["score2"]
                        cell.value = f"{s2}:{s1}"
                    else:
                        cell.value = "—"

            # ── Колонки статистики ──
            r = rank_by_pos.get(i, {})
            pts_cell = ws.cell(row=row_num, column=stats_start_col, value=r.get("points", 0))
            pts_cell.alignment = center
            pts_cell.border = border

            ratio_str = f"{r.get('wins_sum', 0)}:{r.get('losses_sum', 0)}"
            ratio_cell = ws.cell(row=row_num, column=stats_start_col + 1, value=ratio_str)
            ratio_cell.alignment = center
            ratio_cell.border = border

            place_cell = ws.cell(row=row_num, column=stats_start_col + 2, value=r.get("place", ""))
            place_cell.alignment = center
            place_cell.border = border

        # Высота строк
        for r in range(1, size + 2):
            ws.row_dimensions[r].height = 30

    # ── Лист «Общий зачёт» ─────────────────────────────────────────────

    ws_overall = wb.create_sheet(title="Общий зачёт")
    headers = ["Место", "Имя", "Таблица", "Место в табл.", "Очки", "Соотн."]
    for col_idx, h in enumerate(headers, 1):
        c = ws_overall.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.alignment = center
        c.border = border

    ws_overall.column_dimensions["A"].width = 8
    ws_overall.column_dimensions["B"].width = 22
    ws_overall.column_dimensions["C"].width = 10
    ws_overall.column_dimensions["D"].width = 14
    ws_overall.column_dimensions["E"].width = 8
    ws_overall.column_dimensions["F"].width = 10

    for row_idx, o in enumerate(overall, 2):
        cells_data = [
            o["overall_place"],
            o["name"],
            o["table"],
            o["table_place"],
            o["points"],
            f"{o['wins_sum']}:{o['losses_sum']}",
        ]
        for col_idx, val in enumerate(cells_data, 1):
            c = ws_overall.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = center
            c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
